"""Heuristics for locating FABLE Calculator scenario controls."""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import get_column_letter, range_boundaries

from fable_pyculator.spec import (
    FABLE_OUTPUT_SURFACE_SHEETS,
    OutputTable,
    ScenarioParameter,
    SelectionControl,
    SelectionOption,
)


SCENARIO_SHEET_HINTS = ("scenario", "scenarios")
INPUT_LABEL_HINTS = ("scenario", "target", "assumption", "parameter", "select", "choice")


def discover_scenario_parameters(
    workbook_path: str | Path,
    *,
    sheet_hints: Iterable[str] = SCENARIO_SHEET_HINTS,
    label_hints: Iterable[str] = INPUT_LABEL_HINTS,
    max_rows: int = 250,
    max_columns: int = 40,
) -> list[ScenarioParameter]:
    """Return likely scenario controls from visible FABLE Calculator sheets.

    This is deliberately heuristic. It finds non-formula cells near text labels on sheets whose names
    look scenario-related; country-specific wrappers should review and curate the result into a
    committed spec before treating it as a stable user interface.
    """

    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    lowered_sheet_hints = tuple(hint.casefold() for hint in sheet_hints)
    lowered_label_hints = tuple(hint.casefold() for hint in label_hints)
    parameters: list[ScenarioParameter] = []

    for worksheet in workbook.worksheets:
        if not any(hint in worksheet.title.casefold() for hint in lowered_sheet_hints):
            continue
        for row in worksheet.iter_rows(max_row=max_rows, max_col=max_columns):
            label_cell = _first_label_cell(row, lowered_label_hints)
            if label_cell is None:
                continue
            label, label_column = label_cell
            value_cells = [cell for cell in row if getattr(cell, "column", 0) > label_column]
            if not value_cells:
                value_cells = list(row)
            for cell in value_cells:
                if _is_editable_value_cell(cell):
                    parameters.append(
                        ScenarioParameter(
                            name=_parameter_name(worksheet.title, cell.coordinate),
                            label=label,
                            cell_ref=f"{worksheet.title}!{cell.coordinate}",
                            kind=_control_kind(cell.value),
                            default=cell.value,
                            source="heuristic",
                        )
                    )
                    break
    return parameters


def discover_selection_controls(
    workbook_path: str | Path,
    *,
    sheet_name: str = "SCENARIOS selection",
) -> list[SelectionControl]:
    """Discover mutually-exclusive ``x`` selection tables on the FABLE scenario sheet."""

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    worksheet = workbook[sheet_name]
    controls: list[SelectionControl] = []
    for table_name in worksheet.tables.keys():
        table = worksheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [worksheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
        if len(headers) < 2 or str(headers[0]).casefold() != "selection":
            continue
        code_header = str(headers[1])
        label = _control_label(worksheet.cell(min_row - 1, min_col + 1).value, table_name)
        location = _optional_text(worksheet.cell(min_row - 1, min_col).value)
        options: list[SelectionOption] = []
        for row in range(min_row + 1, max_row + 1):
            value = worksheet.cell(row, min_col + 1).value
            if value is None:
                continue
            selected_marker = worksheet.cell(row, min_col).value
            options.append(
                SelectionOption(
                    value=str(value),
                    label=str(value),
                    selection_cell_ref=f"{worksheet.title}!{worksheet.cell(row, min_col).coordinate}",
                    description=_optional_text(worksheet.cell(row, min_col + 2).value),
                    selected=isinstance(selected_marker, str) and selected_marker.strip().casefold() == "x",
                )
            )
        controls.append(
            SelectionControl(
                name=_parameter_name("", table_name),
                label=label,
                table_name=table_name,
                sheet=worksheet.title,
                range_ref=table.ref,
                code_header=code_header,
                options=options,
                location=location,
            )
        )
    return sorted(controls, key=lambda control: _location_sort_key(control.location, control.table_name))


def discover_output_tables(
    workbook_path: str | Path,
    *,
    sheet_names: Iterable[str] = FABLE_OUTPUT_SURFACE_SHEETS,
) -> list[OutputTable]:
    """Discover Excel tables on the canonical FABLE output data sheets."""

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    tables: list[OutputTable] = []
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        for table_name in worksheet.tables.keys():
            table = worksheet.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            column_labels = tuple(
                _column_label(worksheet.cell(min_row, column).value, column)
                for column in range(min_col, max_col + 1)
            )
            row_labels = tuple(
                _row_label(worksheet.cell(row, min_col).value, row)
                for row in range(min_row + 1, max_row + 1)
            )
            cell_refs = tuple(
                tuple(
                    f"{worksheet.title}!{get_column_letter(column)}{row}"
                    for column in range(min_col, max_col + 1)
                )
                for row in range(min_row + 1, max_row + 1)
            )
            tables.append(
                OutputTable(
                    name=_parameter_name(sheet_name, table_name),
                    sheet=worksheet.title,
                    range_ref=table.ref,
                    cell_refs=cell_refs,
                    row_labels=row_labels,
                    column_labels=column_labels,
                    label=table_name,
                )
            )
    return tables


def _first_label_cell(row: Iterable[Cell], label_hints: tuple[str, ...]) -> tuple[str, int] | None:
    for cell in row:
        if isinstance(cell.value, str):
            text = " ".join(cell.value.split())
            if text and any(hint in text.casefold() for hint in label_hints):
                return text, cell.column
    return None


def _is_editable_value_cell(cell: Cell) -> bool:
    if cell.value is None:
        return False
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return False
    return isinstance(cell.value, str | int | float | bool)


def _control_kind(value: object) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int | float) and not isinstance(value, bool):
        return "number"
    return "text"


def _parameter_name(sheet: str, coordinate: str) -> str:
    stem = "".join(character.lower() if character.isalnum() else "_" for character in sheet)
    suffix = "".join(character.lower() if character.isalnum() else "_" for character in coordinate)
    name = "_".join(part for part in f"{stem}_{suffix}".split("_") if part)
    return name


def _control_label(value: object, fallback: str) -> str:
    text = _optional_text(value)
    return text or fallback


def _optional_text(value: object) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).split())
    return text or None


def _location_sort_key(location: str | None, fallback: str) -> tuple[int, str]:
    if location and location.startswith("S."):
        try:
            return int(location.removeprefix("S.").rstrip(".")), fallback
        except ValueError:
            pass
    return 999, fallback


def _column_label(value: object, column: int) -> str:
    return _optional_text(value) or get_column_letter(column)


def _row_label(value: object, row: int) -> str:
    return _optional_text(value) or str(row)
