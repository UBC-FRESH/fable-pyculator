from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from fable_pyculator import (
    FableCalculatorSpec,
    HeadlinePoint,
    HeadlineSeries,
    OutputTable,
    build_cached_workbook_validation_scenario,
    build_modelwright_freshforge_workflow,
    derive_output_refs,
    derive_output_refs_for_strategy,
    fable_freshforge_build_paths,
    freshforge_2021_build_paths,
    prepare_freshforge_rebuild,
    prepare_2021_freshforge_rebuild,
    write_freshforge_workflow,
    write_output_refs,
    write_validation_scenario,
)


def test_derive_output_refs_filters_by_wildcard_and_table_name() -> None:
    spec = _spec()

    assert derive_output_refs(spec) == ("GHG!B3", "GHG!D3", "LAND!B3")
    assert derive_output_refs(spec, column_flavour_tags="OUTPUT", table_names=("ghg_resultsghg",)) == (
        "GHG!B3",
        "GHG!D3",
    )
    assert derive_output_refs(spec, column_flavour_tags="DATA*") == ("GHG!C3",)


def test_fable_freshforge_build_paths_are_version_general(tmp_path: Path) -> None:
    paths_2020 = fable_freshforge_build_paths(workbook_version="2020", repo_root=tmp_path)
    paths_2021 = fable_freshforge_build_paths(workbook_version="2021", repo_root=tmp_path)
    paths_2022 = fable_freshforge_build_paths(workbook_version="2022", repo_root=tmp_path)

    assert paths_2020.workbook_path == tmp_path / "tmp/private-workbooks/2020_Open_FABLECalculator.xlsx"
    assert paths_2020.artifact_dir == tmp_path / "tmp/generated-models/fable-2020"
    assert paths_2020.generated_model_path.name == "generated_fable_2020_model.py"
    assert paths_2021 == freshforge_2021_build_paths(repo_root=tmp_path)
    assert paths_2022.generated_model_path == tmp_path / "tmp/generated-models/fable-2022/generated_fable_2022_model.py"


def test_derive_output_refs_for_named_strategies() -> None:
    spec = _spec()

    assert derive_output_refs_for_strategy(spec, strategy="output-columns") == ("GHG!B3", "GHG!D3", "LAND!B3")
    assert derive_output_refs_for_strategy(spec, strategy="headline-only") == ("GHG!B3", "LAND!B3")
    assert derive_output_refs_for_strategy(
        spec,
        strategy="table",
        table_names=("ghg_resultsghg",),
    ) == ("GHG!B3", "GHG!D3")
    assert derive_output_refs_for_strategy(spec, strategy="flavour-tags", column_flavour_tags="DATA*") == ("GHG!C3",)
    assert derive_output_refs_for_strategy(spec, strategy="all-columns", table_names=("land_resultsland",)) == (
        "LAND!A3",
        "LAND!B3",
    )


def test_derive_output_refs_for_strategy_rejects_invalid_requests() -> None:
    spec = _spec()

    try:
        derive_output_refs_for_strategy(spec, strategy="table")
    except ValueError as error:
        assert "requires at least one table name" in str(error)
    else:  # pragma: no cover - defensive assertion
        raise AssertionError("expected ValueError")

    try:
        derive_output_refs_for_strategy(spec, strategy="missing")  # type: ignore[arg-type]
    except ValueError as error:
        assert "unknown output-ref strategy" in str(error)
    else:  # pragma: no cover - defensive assertion
        raise AssertionError("expected ValueError")


def test_derive_output_refs_rejects_unknown_table_name() -> None:
    try:
        derive_output_refs(_spec(), table_names=("missing",))
    except KeyError as error:
        assert "missing" in str(error)
    else:  # pragma: no cover - defensive assertion
        raise AssertionError("expected KeyError")


def test_write_output_refs_writes_sorted_unique_json(tmp_path: Path) -> None:
    output_path = tmp_path / "nested" / "output_refs.json"

    written = write_output_refs(output_path, ["GHG!D3", "GHG!B3", "GHG!B3"])

    assert written == ("GHG!B3", "GHG!D3")
    assert json.loads(output_path.read_text(encoding="utf-8")) == ["GHG!B3", "GHG!D3"]


def test_validation_scenario_uses_cached_workbook_values(tmp_path: Path) -> None:
    workbook_path = tmp_path / "cached.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GHG"
    sheet["B3"] = 42
    sheet["C3"] = None
    sheet["D3"] = "#N/A"
    workbook.save(workbook_path)

    scenario = build_cached_workbook_validation_scenario(
        workbook_path,
        ["GHG!B3", "GHG!C3", "GHG!D3"],
        generated_model_path="tmp/generated-models/fable-2021/generated_fable_2021_model.py",
        scenario_id="test-scenario",
        description="Synthetic validation scenario.",
    )

    assert scenario["scenario_id"] == "test-scenario"
    assert scenario["outputs"] == [
        {"cell_ref": "GHG!B3", "kind": "number", "tolerance": 1e-9},
        {"cell_ref": "GHG!D3", "kind": "error"},
    ]


def test_workflow_builder_declares_modelwright_nodes_and_relative_artifacts(tmp_path: Path) -> None:
    paths = freshforge_2021_build_paths(repo_root=tmp_path)

    workflow = build_modelwright_freshforge_workflow(
        paths,
        workdir=tmp_path,
        workflow_id="fable_2021_modelwright_run",
        name="FABLE 2021 Modelwright run",
        description="Synthetic workflow.",
        module_name="generated_fable_2021_model",
    )

    assert [node["id"] for node in workflow["nodes"]] == [
        "infer_contract",
        "generate_model",
        "execute_model",
        "evaluate_model",
    ]
    assert workflow["nodes"][0]["parameters"]["workbook"] == "tmp/private-workbooks/2021_Open_FABLECalculator.xlsx"
    assert workflow["nodes"][0]["artifacts"]["output_refs"] == "tmp/generated-models/fable-2021/output_refs.json"
    assert workflow["nodes"][1]["artifacts"]["generated_model"] == (
        "tmp/generated-models/fable-2021/generated_fable_2021_model.py"
    )
    assert workflow["nodes"][3]["artifacts"]["evaluation_report"] == (
        "tmp/generated-models/fable-2021/evaluation-report.json"
    )


def test_workflow_and_validation_json_writers_are_stable(tmp_path: Path) -> None:
    workflow_path = tmp_path / "workflow.json"
    scenario_path = tmp_path / "validation-scenario.json"
    workflow = {"workflow": {"id": "demo"}, "nodes": []}
    scenario = {"scenario_id": "demo", "outputs": []}

    assert write_freshforge_workflow(workflow_path, workflow) == workflow
    assert write_validation_scenario(scenario_path, scenario) == scenario

    assert workflow_path.read_text(encoding="utf-8").endswith("\n")
    assert scenario_path.read_text(encoding="utf-8").endswith("\n")
    assert json.loads(workflow_path.read_text(encoding="utf-8")) == workflow
    assert json.loads(scenario_path.read_text(encoding="utf-8")) == scenario


def test_prepare_2021_freshforge_rebuild_writes_plan_artifacts(tmp_path: Path) -> None:
    workbook_path = tmp_path / "tmp" / "private-workbooks" / "2021_Open_FABLECalculator.xlsx"
    workbook_path.parent.mkdir(parents=True)
    workbook = Workbook()
    ghg = workbook.active
    ghg.title = "GHG"
    ghg["B3"] = 42
    ghg["D3"] = 84
    land = workbook.create_sheet("LAND")
    land["B3"] = 12
    workbook.save(workbook_path)

    plan = prepare_2021_freshforge_rebuild(repo_root=tmp_path, spec=_spec())

    assert plan.output_refs == ("GHG!B3", "GHG!D3", "LAND!B3")
    assert plan.comparable_output_count == 3
    assert json.loads(plan.paths.output_refs_path.read_text(encoding="utf-8")) == [
        "GHG!B3",
        "GHG!D3",
        "LAND!B3",
    ]
    assert plan.paths.validation_scenario_path.exists()
    assert plan.paths.workflow_path.exists()
    assert plan.workflow["nodes"][0]["provider"] == "modelwright.model_infer_contract"


def test_prepare_freshforge_rebuild_uses_version_specific_defaults(tmp_path: Path) -> None:
    workbook_path = tmp_path / "tmp" / "private-workbooks" / "2022_Open_FABLECalculator.xlsx"
    workbook_path.parent.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GHG"
    sheet["B3"] = 42
    sheet["D3"] = 84
    land = workbook.create_sheet("LAND")
    land["B3"] = 12
    workbook.save(workbook_path)

    plan = prepare_freshforge_rebuild(workbook_version="2022", repo_root=tmp_path, spec=_spec())

    assert plan.paths.workflow_path == tmp_path / "tmp/generated-models/fable-2022/freshforge-modelwright-run-workflow.json"
    assert plan.paths.generated_model_path.name == "generated_fable_2022_model.py"
    assert plan.workflow["nodes"][0]["parameters"]["module_name"] == "generated_fable_2022_model"
    assert plan.validation_scenario["scenario_id"] == "fable-c-2022-freshforge-rebuild"


def _spec() -> FableCalculatorSpec:
    return FableCalculatorSpec(
        output_tables=(
            OutputTable(
                name="ghg_resultsghg",
                sheet="GHG",
                range_ref="A2:D3",
                cell_refs=(("GHG!A3", "GHG!B3", "GHG!C3", "GHG!D3"),),
                row_labels=("2030",),
                column_labels=("Year", "TotalCO2e", "Data", "OtherOutput"),
                column_flavour_tags=("DIRECT", "OUTPUT - 8", "DATA-1", "OUTPUT-9"),
            ),
            OutputTable(
                name="land_resultsland",
                sheet="LAND",
                range_ref="A2:B3",
                cell_refs=(("LAND!A3", "LAND!B3"),),
                row_labels=("2030",),
                column_labels=("Year", "Area"),
                column_flavour_tags=("DIRECT", "OUTPUT-4"),
            ),
        ),
        headline_series=(
            HeadlineSeries(
                name="ghg_total_co2e",
                label="Total GHG emissions",
                group="GHG",
                sheet="GHG",
                table_name="ResultsGHG",
                points=(HeadlinePoint(year=2030, cell_refs=("GHG!B3",)),),
            ),
            HeadlineSeries(
                name="land_total_area",
                label="Total land area",
                group="LAND",
                sheet="LAND",
                table_name="ResultsLand",
                points=(HeadlinePoint(year=2030, cell_refs=("LAND!B3",)),),
            ),
        ),
    )
