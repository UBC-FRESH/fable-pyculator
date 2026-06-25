from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from fable_pyculator import discover_output_tables, discover_scenario_parameters, discover_selection_controls


def test_discover_scenario_parameters_finds_labeled_values(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SCENARIOS selection"
    worksheet["A1"] = "Scenario ambition"
    worksheet["D1"] = 2
    worksheet["A2"] = "Formula output"
    worksheet["D2"] = "=D1*2"
    path = tmp_path / "fable.xlsx"
    workbook.save(path)

    parameters = discover_scenario_parameters(path)

    assert len(parameters) == 1
    assert parameters[0].cell_ref == "SCENARIOS selection!D1"
    assert parameters[0].default == 2


def test_discover_selection_controls_finds_x_marker_table(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SCENARIOS selection"
    worksheet["A1"] = "S.1"
    worksheet["B1"] = "GDP projections"
    worksheet.append(["SELECTION", "GDP_SCEN", "DESCRIPTION"])
    worksheet.append([None, "SSP1", "Sustainability"])
    worksheet.append(["x", "SSP2", "Middle of the road"])
    worksheet.append([None, "SSP3", "Fragmentation"])
    worksheet.add_table(Table(displayName="GDP_Scen", ref="A2:C5"))
    path = tmp_path / "fable.xlsx"
    workbook.save(path)

    controls = discover_selection_controls(path)

    assert len(controls) == 1
    control = controls[0]
    assert control.name == "gdp_scen"
    assert control.label == "GDP projections"
    assert control.location == "S.1"
    assert control.default == "SSP2"
    assert control.input_mapping("SSP3") == {
        "SCENARIOS selection!A3": None,
        "SCENARIOS selection!A4": None,
        "SCENARIOS selection!A5": "x",
    }


def test_discover_output_tables_finds_tables_on_canonical_output_sheets(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FOOD"
    worksheet.append(["Metric", "2020", "2030"])
    worksheet.append(["Calories", 2500, 2600])
    worksheet.append(["Protein", 80, 82])
    worksheet.add_table(Table(displayName="Results_Diets", ref="A1:C3"))
    path = tmp_path / "fable.xlsx"
    workbook.save(path)

    tables = discover_output_tables(path)

    assert len(tables) == 1
    table = tables[0]
    assert table.name == "food_results_diets"
    assert table.sheet == "FOOD"
    assert table.range_ref == "A1:C3"
    assert table.column_labels == ("Metric", "2020", "2030")
    assert table.row_labels == ("Calories", "Protein")
    assert table.cell_refs == (
        ("FOOD!A2", "FOOD!B2", "FOOD!C2"),
        ("FOOD!A3", "FOOD!B3", "FOOD!C3"),
    )
