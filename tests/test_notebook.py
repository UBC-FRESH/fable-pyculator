from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from fable_pyculator import (
    build_2020_notebook_spec,
    load_generated_model,
    run_2020_notebook_loop,
    run_notebook_loop,
)


def test_load_generated_model_imports_calculate_module(tmp_path: Path) -> None:
    model_path = tmp_path / "generated_fable.py"
    model_path.write_text(
        "def calculate(inputs=None):\n"
        "    return {'GHG!B2': 42, 'input_count': len(inputs or {})}\n",
        encoding="utf-8",
    )

    module = load_generated_model(model_path, module_name="test_generated_fable")

    assert module.calculate({"SCENARIOS selection!A3": "x"}) == {"GHG!B2": 42, "input_count": 1}


def test_load_generated_model_reports_missing_path(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError, match="generated model not found"):
        load_generated_model(tmp_path / "missing.py")


def test_run_notebook_loop_applies_selection_controls_and_renders_artifacts() -> None:
    spec = build_2020_notebook_spec(_synthetic_workbook_path())

    def calculate(inputs=None):
        assert inputs == {
            "SCENARIOS selection!A3": "x",
            "SCENARIOS selection!A4": None,
        }
        return {
            "GHG!A2": 2030,
            "GHG!B2": 42,
            "FOOD!C2": 2500,
            "LAND!B2": 100,
            "WATER!C2": 1,
            "WATER!D2": 2,
            "WATER!E2": 3,
            "WATER!F2": 4,
            "WATER!G2": 5,
            "WATER!H2": 6,
        }

    result = run_notebook_loop(
        calculate,
        spec,
        {"gdp_scen": "SSP1"},
        output_table_names=("ghg_resultsghg",),
        headline_series_names=("ghg_total_co2e", "water_total_footprint"),
        include_figures=False,
    )

    assert result.run.inputs == {
        "SCENARIOS selection!A3": "x",
        "SCENARIOS selection!A4": None,
    }
    assert result.output_tables["ghg_resultsghg"].loc["2030", "TotalCO2e"] == 42
    assert result.headline_frames["ghg_total_co2e"].loc[0, "value"] == 42
    assert result.headline_frames["water_total_footprint"].loc[0, "value"] == 21
    assert result.headline_figures == {}


def test_run_2020_notebook_loop_loads_ignored_model_path(tmp_path: Path) -> None:
    workbook_path = _synthetic_workbook_path(tmp_path / "synthetic_fable.xlsx")
    model_path = tmp_path / "generated_fable_2020_model.py"
    model_path.write_text(
        "def calculate(inputs=None):\n"
        "    return {\n"
        "        'GHG!A2': 2030,\n"
        "        'GHG!B2': 42,\n"
        "        'FOOD!C2': 2500,\n"
        "        'LAND!B2': 100,\n"
        "        'WATER!C2': 1,\n"
        "        'WATER!D2': 2,\n"
        "        'WATER!E2': 3,\n"
        "        'WATER!F2': 4,\n"
        "        'WATER!G2': 5,\n"
        "        'WATER!H2': 6,\n"
        "    }\n",
        encoding="utf-8",
    )

    result = run_2020_notebook_loop(
        {"gdp_scen": "SSP2"},
        workbook_path=workbook_path,
        generated_model_path=model_path,
        output_table_names=("ghg_resultsghg",),
        headline_series_names=("ghg_total_co2e",),
        include_figures=False,
    )

    assert result.run.inputs == {
        "SCENARIOS selection!A3": None,
        "SCENARIOS selection!A4": "x",
    }
    assert result.headline_frames["ghg_total_co2e"].loc[0, "value"] == 42


def _synthetic_workbook_path(path: Path | None = None) -> Path:
    workbook = Workbook()
    scenarios = workbook.active
    scenarios.title = "SCENARIOS selection"
    scenarios["A1"] = "S.1"
    scenarios["B1"] = "GDP projections"
    scenarios.append(["SELECTION", "GDP_SCEN", "DESCRIPTION"])
    scenarios.append([None, "SSP1", "Sustainability"])
    scenarios.append(["x", "SSP2", "Middle of the road"])
    scenarios.add_table(Table(displayName="GDP_Scen", ref="A2:C4"))

    indextables = workbook.create_sheet("Indextables")
    indextables.append(["Table", "Description"])
    indextables.append(["Total_Results_diets", "Diet totals"])
    indextables.append(["ResultsLand", "Land totals"])
    indextables.append(["ResultsGHG", "GHG totals"])
    indextables.append(["TotalResultsWF", "Water totals"])

    food = workbook.create_sheet("FOOD")
    food.append(["PROD_GROUP", "YEAR", "kcal_feas"])
    food.append(["TOTAL", 2030, 2500])
    food.add_table(Table(displayName="Total_results_diets", ref="A1:C2"))

    land = workbook.create_sheet("LAND")
    land.append(["Year", "TOTAL"])
    land.append([2030, 100])
    land.add_table(Table(displayName="ResultsLand", ref="A1:B2"))

    ghg = workbook.create_sheet("GHG")
    ghg.append(["Year", "TotalCO2e"])
    ghg.append([2030, 42])
    ghg.add_table(Table(displayName="ResultsGHG", ref="A1:B2"))

    water = workbook.create_sheet("WATER")
    water.append(
        [
            "Product",
            "YEAR",
            "wf_green_crop",
            "wf_blue_crop",
            "wf_grey_crop",
            "wf_green_live",
            "wf_blue_live",
            "wf_grey_live",
        ]
    )
    water.append(["TOTAL", 2030, 1, 2, 3, 4, 5, 6])
    water.add_table(Table(displayName="TotalResultsWF", ref="A1:H2"))

    target = path or Path("tmp/test-scratch/synthetic_fable.xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target
